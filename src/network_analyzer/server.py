"""Live Smith chart web viewer for the Agilent E5061B ENA.

Run:
    uvicorn network_analyzer.server:app --host 0.0.0.0 --port 8766

Open http://localhost:8766 in a browser.
"""

from __future__ import annotations

import asyncio
import json
import os
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.styles import Border, Side
from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from network_analyzer import NetworkAnalyzer

STATIC_DIR = Path(__file__).parent / "static"

# ---------------------------------------------------------------------------
# shared state
# ---------------------------------------------------------------------------

na = NetworkAnalyzer()
sweep_lock = asyncio.Lock()


@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        na.connect()
    except Exception as e:
        print(f"[startup] ENA auto-connect failed: {e}. Use /api/connect to retry.")
    yield
    try:
        na.disconnect()
    except Exception:
        pass


app = FastAPI(title="ENA Smith Live Viewer", lifespan=lifespan)


# ---------------------------------------------------------------------------
# REST: connect / disconnect / status
# ---------------------------------------------------------------------------

@app.get("/api/status")
def get_status():
    if not na.is_connected:
        return {"connected": False, "idn": None}
    try:
        return {"connected": True, "idn": na.idn}
    except Exception as e:
        return {"connected": False, "idn": None, "error": str(e)}


@app.post("/api/connect")
async def api_connect():
    if na.is_connected:
        return {"connected": True, "idn": na.idn}
    try:
        await asyncio.to_thread(na.connect)
        return {"connected": True, "idn": na.idn}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Connect failed: {e}")


@app.post("/api/disconnect")
async def api_disconnect():
    try:
        await asyncio.to_thread(na.disconnect)
        return {"connected": False}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Disconnect failed: {e}")


# ---------------------------------------------------------------------------
# config models
# ---------------------------------------------------------------------------

class SweepConfig(BaseModel):
    start_freq_hz: Optional[float] = None
    stop_freq_hz:  Optional[float] = None
    points:        Optional[int]   = None
    if_bandwidth_hz: Optional[float] = None
    source_power_dbm: Optional[float] = None


# ---------------------------------------------------------------------------
# REST: config
# ---------------------------------------------------------------------------

@app.get("/api/config")
def get_config():
    return {
        "idn": na.idn,
        "start_freq_hz": na.start_freq_hz,
        "stop_freq_hz":  na.stop_freq_hz,
        "points":        na.points,
        "if_bandwidth_hz": na.if_bandwidth_hz,
        "source_power_dbm": na.source_power_dbm,
    }


@app.post("/api/config")
async def set_config(cfg: SweepConfig):
    async with sweep_lock:
        if cfg.start_freq_hz   is not None: na.start_freq_hz   = cfg.start_freq_hz
        if cfg.stop_freq_hz    is not None: na.stop_freq_hz    = cfg.stop_freq_hz
        if cfg.points          is not None: na.points          = cfg.points
        if cfg.if_bandwidth_hz is not None: na.if_bandwidth_hz = cfg.if_bandwidth_hz
        if cfg.source_power_dbm is not None: na.source_power_dbm = cfg.source_power_dbm
    return get_config()


# ---------------------------------------------------------------------------
# WebSocket: live sweep stream
# ---------------------------------------------------------------------------

@app.websocket("/ws/sweep")
async def ws_sweep(ws: WebSocket):
    await ws.accept()
    try:
        while True:
            async with sweep_lock:
                # blocking ENA calls in a worker thread so the event loop stays free
                freqs, gamma = await asyncio.to_thread(_sweep_and_read)

            r = (50.0 * (1 + gamma) / (1 - gamma)).real
            x = (50.0 * (1 + gamma) / (1 - gamma)).imag
            payload = {
                "freqs": freqs.tolist(),
                "gamma_real": gamma.real.tolist(),
                "gamma_imag": gamma.imag.tolist(),
                "r": r.tolist(),
                "x": x.tolist(),
            }
            await ws.send_text(json.dumps(payload))
    except WebSocketDisconnect:
        return


def _sweep_and_read():
    na.sweep()
    return na.get_s_parameter_complex("S11")


# ---------------------------------------------------------------------------
# REST: measure at specific frequencies
# ---------------------------------------------------------------------------

class MeasureRequest(BaseModel):
    frequencies_hz: Optional[list[float]] = None  # None = return all ENA sweep points
    parameter: str = "S11"
    z0: float = 50.0


@app.post("/api/measure")
async def measure_at_frequencies(req: MeasureRequest):
    async with sweep_lock:
        result = await asyncio.to_thread(
            _measure_freqs, req.frequencies_hz, req.parameter, req.z0
        )
    return result


def _measure_freqs(frequencies_hz: Optional[list[float]], parameter: str, z0: float) -> list[dict]:
    # Read current live data — do not trigger a sweep or change any ENA settings
    sweep_freqs, gamma = na.get_s_parameter_complex(parameter)

    if frequencies_hz:
        freqs_arr = np.array(sorted(set(frequencies_hz)))
        gamma_r = np.interp(freqs_arr, sweep_freqs, gamma.real)
        gamma_i = np.interp(freqs_arr, sweep_freqs, gamma.imag)
    else:
        freqs_arr = sweep_freqs
        gamma_r = gamma.real
        gamma_i = gamma.imag

    gamma_c = gamma_r + 1j * gamma_i
    z = z0 * (1 + gamma_c) / (1 - gamma_c)
    s_db = 20.0 * np.log10(np.abs(gamma_c))
    l_nh = (z.imag / (2 * np.pi * freqs_arr)) * 1e9

    return [
        {
            "freq_hz": float(f),
            "gamma_real": round(float(gr), 6),
            "gamma_imag": round(float(gi), 6),
            "r_ohm": round(float(r), 4),
            "x_ohm": round(float(x), 4),
            "l_nh": round(float(l), 4),
            "s11_db": round(float(db), 4),
        }
        for f, gr, gi, r, x, l, db in zip(
            freqs_arr, gamma_r, gamma_i, z.real, z.imag, l_nh, s_db
        )
    ]


# ---------------------------------------------------------------------------
# REST: append results to Excel
# ---------------------------------------------------------------------------

HEADERS = ["Timestamp", "Freq (MHz)", "R (Ohm)", "X (Ohm)", "L (nH)", "S11 (dB)"]

class AppendResultsRequest(BaseModel):
    file_path: str
    sheet_name: str = "Results"
    rows: list[dict]  # each dict must have the HEADERS keys


@app.post("/api/append-results")
async def append_results(req: AppendResultsRequest):
    path = Path(req.file_path)
    await asyncio.to_thread(_append_excel, path, req.sheet_name, req.rows)
    return {"status": "ok", "appended": len(req.rows), "sheet": req.sheet_name}


def _append_excel(path: Path, sheet_name: str, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
    else:
        ws = wb[sheet_name]

    for row in rows:
        ws.append([row.get(h) for h in HEADERS])

    # bottom border on last row of this run
    bottom = Border(bottom=Side(style='medium'))
    last_row = ws.max_row
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=last_row, column=col).border = bottom

    wb.save(path)


# ---------------------------------------------------------------------------
# REST: open a file with its default application
# ---------------------------------------------------------------------------

class OpenFileRequest(BaseModel):
    file_path: str


@app.post("/api/open-file")
async def open_file(req: OpenFileRequest):
    path = Path(req.file_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {req.file_path}")
    await asyncio.to_thread(os.startfile, str(path))
    return {"status": "ok", "opened": str(path)}


# ---------------------------------------------------------------------------
# static files
# ---------------------------------------------------------------------------

@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
